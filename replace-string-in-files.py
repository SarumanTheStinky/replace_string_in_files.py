"""
replace_string_in_files.py

Recursively scans all files in the current directory and subdirectories.
Replaces all occurrences of a target string with a new string in:
- .xlsx files (cell contents)
- .xml files (UTF-8)
- all other UTF-8-readable text files

Optional: exclude files or file extensions, enable logging for debug.
"""

import os
import openpyxl
import codecs

# === CONFIGURATION ===
renamefolderpath = os.path.dirname(__file__)  # Current directory
oldString = "SarumanTheStupid"
newString = "SarumanTheStinky"

def transform_string_in_files(root_dir, oldString, newString
    # === OPTIONAL: exclude specific file extensions ===
    # , except_file_extensions=(
    #     ".tiff",
    #     ".czi",
    #     ".jpg",
    #     ".html",
    #     ".db",
    #     ".jpeg",
    #     ".py",
    #     ".pdf",
    #     ".png",
    #     ".pptx",
    #     ".otf",
    #     ".bin",
    #     ".tif"
    # )
):
    for root, dirs, files in os.walk(root_dir):
        for file in files:

            # === OPTIONAL: exclude specific filenames or file endings ===
            # if not (file.endswith(except_file_extensions) or file in ("legollum.xlsx", "gimlum.xlsx")):

            file_path = os.path.join(root, file)

            # === DEBUG: print file path being processed ===
            # print("Processing file:", file_path)

            try:
                if file.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(file_path)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value and oldString in str(cell.value):
                                    cell.value = str(cell.value).replace(oldString, newString)
                                    if cell.value == oldString:
                                        raise Exception(f"Failed to replace string '{oldString}' in {file_path}")
                    wb.save(file_path)

                elif file.endswith('.xml'):
                    with codecs.open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    if oldString in content:
                        print("Original string found in file: ", file_path)
                        updated_content = content.replace(oldString, newString)
                        with codecs.open(file_path, 'w', encoding='utf-8', errors='ignore') as f:
                            f.write(updated_content)
                        print("File updated.")

                    # === DEBUG: Uncomment to log if string not found ===
                    # else:
                    #     print("Original string not found in file.")

                else:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    if oldString in content:
                        print("Original string found in file: ", file_path)
                        updated_content = content.replace(oldString, newString)
                        with open(file_path, 'w', encoding='utf-8') as f:
                            f.write(updated_content)
                        print("File updated.")

                    # === DEBUG: Uncomment to log if string not found ===
                    # else:
                    #     print("Original string not found in file.")

            except KeyError as e:
                print("Error processing file: ", file_path)
                print("Error message:", str(e))
                raise Exception(f"Missing drawing in {file_path}")

# === MAIN ===
transform_string_in_files(renamefolderpath, oldString, newString)
